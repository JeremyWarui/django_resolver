"""Report generation views — GET /reports/types/ and GET /reports/generate/.

Scope is derived from the JWT role claim (same pattern as analytics views).
All reports use scoped_ticket_qs so every role only exports their own data.
Every report includes a Summary sheet that mirrors the analytics overview cards
(open_backlog, SLA compliance, CSAT, resolution p50/p90) so the Excel data
is consistent with what the analytics/reports pages show in the UI.
"""

from collections import defaultdict
from datetime import timedelta
from io import BytesIO

import openpyxl
from django.db.models import Count, F, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication

from apps.analytics.services import ACTIVE_STATUSES, aggregate, resolve_date_range
from apps.common.roles import resolve_role
from apps.tickets.models import Ticket
from apps.tickets.services.scope import scoped_ticket_qs

# ── Excel styling ──────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_headers(ws, num_cols: int) -> None:
    ws.row_dimensions[1].height = 20
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 2, 50
        )


def _fmt(dt) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


# ── Date-range and queryset helpers ───────────────────────────────────────────


def _build_date_range_params(request) -> dict:
    """Build params dict compatible with resolve_date_range()."""
    params: dict = {}
    start = request.query_params.get("start_date")
    end = request.query_params.get("end_date")
    timeframe = request.query_params.get("timeframe", "all")

    if start and end:
        params["date_from"] = start
        params["date_to"] = end
    else:
        offsets = {"day": 1, "week": 7, "month": 30, "quarter": 90, "year": 365}
        if timeframe in offsets:
            params["days"] = offsets[timeframe]
        # 'all' → no date filter; resolve_date_range defaults to 30d,
        # so for 'all' we just skip the date filter below.

    return params


def _base_qs(request, role):
    """Return scoped + date-filtered queryset; 'all time' returns unfiltered scope."""
    if role:
        qs = scoped_ticket_qs(request.user, role)
    else:
        qs = Ticket.objects.filter(raised_by=request.user)

    timeframe = request.query_params.get("timeframe", "all")
    start = request.query_params.get("start_date")
    end = request.query_params.get("end_date")

    if start and end:
        from datetime import datetime

        try:
            qs = qs.filter(created_at__gte=datetime.strptime(start, "%Y-%m-%d"))
            qs = qs.filter(created_at__lte=datetime.strptime(end, "%Y-%m-%d"))
        except ValueError:
            pass
    elif timeframe != "all":
        offsets = {"day": 1, "week": 7, "month": 30, "quarter": 90, "year": 365}
        if timeframe in offsets:
            since = timezone.now() - timedelta(days=offsets[timeframe])
            qs = qs.filter(created_at__gte=since)

    section_id = request.query_params.get("section_id")
    technician_id = request.query_params.get("technician_id")
    if section_id:
        qs = qs.filter(section_id=section_id)
    if technician_id:
        qs = qs.filter(assigned_to_id=technician_id)

    return qs


def _report_date_range(request, scoped_qs):
    """The window the Summary sheet should use — the SAME effective window as the
    data sheets (_base_qs). A specific timeframe / start-end is honoured directly;
    'all time' spans all scoped data so the Summary matches the all-time sheets
    (no hidden 30-day override)."""
    dr_params = _build_date_range_params(request)
    if dr_params:
        return resolve_date_range(dr_params)

    earliest = (
        scoped_qs.order_by("created_at").values_list("created_at", flat=True).first()
    )
    date_from = earliest or (timezone.now() - timedelta(days=30))
    return resolve_date_range(
        {
            "date_from": date_from.isoformat(),
            "date_to": timezone.now().isoformat(),
        }
    )


# ── Summary sheet (mirrors analytics overview cards) ──────────────────────────


def _sheet_summary(ws, scoped_qs, request) -> None:
    """Summary tab — same numbers as the analytics/reports overview cards.

    Uses the aggregate() core from analytics.services so the data matches
    exactly what the UI shows, over the SAME window as the report's data sheets
    (a specific timeframe is honoured; 'all time' spans all data — no hidden
    30-day override)."""
    date_range = _report_date_range(request, scoped_qs)
    data = aggregate(scoped_qs, date_range)

    _SUMMARY_FILL = PatternFill("solid", fgColor="1E3A5F")
    _LABEL_FILL = PatternFill("solid", fgColor="EFF6FF")

    def _kv(label, value, row):
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.fill = _LABEL_FILL
        label_cell.font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    # Header
    title_cell = ws.cell(row=1, column=1, value="Analytics Summary")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = _SUMMARY_FILL
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    dr = date_range
    ws.cell(row=2, column=1, value="Date Range")
    ws.cell(
        row=2,
        column=2,
        value=f"{dr['date_from'].strftime('%Y-%m-%d')} → {dr['date_to'].strftime('%Y-%m-%d')}",
    )

    row = 4
    _kv("Open Backlog (live)", data.get("open_backlog", 0), row)
    row += 1
    _kv("Created in window", data.get("created", 0), row)
    row += 1
    _kv("Resolved in window", data.get("resolved", 0), row)
    row += 1
    _kv("Net Flow (created − resolved)", data.get("net_flow", 0), row)
    row += 1

    row += 1
    _kv(
        "Resolution SLA %",
        (
            f"{data['resolution_sla_pct']:.1f}%"
            if data.get("resolution_sla_pct") is not None
            else "—"
        ),
        row,
    )
    row += 1
    _kv(
        "Response SLA %",
        (
            f"{data['response_sla_pct']:.1f}%"
            if data.get("response_sla_pct") is not None
            else "—"
        ),
        row,
    )
    row += 1
    _kv("At-Risk tickets", data.get("at_risk", 0), row)
    row += 1
    _kv("Breached tickets", data.get("breached", 0), row)
    row += 1

    row += 1
    _kv("CSAT", f"{data['csat']:.1f}%" if data.get("csat") is not None else "—", row)
    row += 1
    _kv(
        "Reopen Rate",
        f"{data['reopen_rate']:.1f}%" if data.get("reopen_rate") is not None else "—",
        row,
    )
    row += 1
    _kv(
        "Escalation Rate",
        (
            f"{data['escalation_rate']:.1f}%"
            if data.get("escalation_rate") is not None
            else "—"
        ),
        row,
    )
    row += 1

    # p50/p90 resolution times — straight from aggregate() so the Summary matches
    # the analytics endpoints exactly (no re-implemented percentile here).
    p50 = data.get("resolution_time_p50_seconds")
    p90 = data.get("resolution_time_p90_seconds")

    def _fmt_secs(s):
        if s is None:
            return "—"
        h = int(s // 3600)
        m = int((s % 3600) // 60)
        return f"{h}h {m}m" if h else f"{m}m"

    row += 1
    _kv("Resolution p50 (median)", _fmt_secs(p50), row)
    row += 1
    _kv("Resolution p90", _fmt_secs(p90), row)
    row += 1


# ── Sheet builders ─────────────────────────────────────────────────────────────


def _sheet_ticket_lifecycle(ws, qs) -> None:
    headers = [
        "Ticket No",
        "Status",
        "Priority",
        "Level",
        "Raised By",
        "Campus",
        "Service Category",
        "Service Item",
        "Section",
        "Assigned To",
        "Description (200 chars)",
        "Response Due",
        "Resolution Due",
        "Resolved At",
        "Closed At",
        "Created At",
        "Updated At",
        "Paused Since",
        "Total Paused (mins)",
    ]
    ws.append(headers)
    _style_headers(ws, len(headers))
    ws.freeze_panes = "A2"

    tickets = qs.select_related(
        "raised_by",
        "requester_campus",
        "service_item__category",
        "section__section_type",
        "section__campus_department__campus",
        "priority",
        "assigned_to",
    )

    for t in tickets:
        pause_mins = (
            int(t.accumulated_pause.total_seconds() / 60) if t.accumulated_pause else 0
        )
        ws.append(
            [
                t.ticket_no,
                t.get_status_display(),
                t.priority.name if t.priority_id else "",
                t.get_current_level_display(),
                (
                    (t.raised_by.get_full_name() or t.raised_by.username)
                    if t.raised_by_id
                    else ""
                ),
                t.requester_campus.name if t.requester_campus_id else "",
                (
                    t.service_item.category.name
                    if t.service_item_id and t.service_item.category_id
                    else ""
                ),
                t.service_item.name if t.service_item_id else "",
                (
                    t.section.section_type.name
                    if t.section_id and t.section.section_type_id
                    else ""
                ),
                (
                    (t.assigned_to.get_full_name() or t.assigned_to.username)
                    if t.assigned_to_id
                    else "Unassigned"
                ),
                (t.description or "")[:200],
                _fmt(t.response_due_at),
                _fmt(t.resolution_due_at),
                _fmt(t.resolved_at),
                _fmt(t.closed_at),
                _fmt(t.created_at),
                _fmt(t.updated_at),
                _fmt(t.paused_at),
                pause_mins,
            ]
        )

    _auto_width(ws)


def _sheet_technician_performance(ws, qs) -> None:
    headers = [
        "Technician",
        "Username",
        "Total Assigned",
        "Open",
        "Resolved",
        "Closed",
        "Pending",
        "Escalated",
        "Avg Resolution (hrs)",
    ]
    ws.append(headers)
    _style_headers(ws, len(headers))
    ws.freeze_panes = "A2"

    by_tech = (
        qs.filter(assigned_to__isnull=False)
        .values(
            "assigned_to",
            "assigned_to__username",
            "assigned_to__first_name",
            "assigned_to__last_name",
        )
        .annotate(
            total=Count("id"),
            open_count=Count("id", filter=Q(status__in=ACTIVE_STATUSES)),
            resolved_count=Count("id", filter=Q(status="resolved")),
            closed_count=Count("id", filter=Q(status="closed")),
            pending_count=Count("id", filter=Q(status="pending")),
            escalated_count=Count("id", filter=~Q(current_level="technician")),
        )
        .order_by("-total")
    )

    # Average resolution time per technician — ONE query, grouped in Python
    # (replaces the previous per-technician query inside the loop / N+1).
    res_hours_by_tech = defaultdict(list)
    for tech_id, resolved_at, created_at, pause in qs.filter(
        assigned_to__isnull=False,
        status__in=["resolved", "closed"],
        resolved_at__isnull=False,
    ).values_list("assigned_to_id", "resolved_at", "created_at", "accumulated_pause"):
        if resolved_at and created_at:
            delta = (resolved_at - created_at) - (pause or timedelta())
            res_hours_by_tech[tech_id].append(max(delta.total_seconds(), 0) / 3600)

    for row in by_tech:
        first = row["assigned_to__first_name"] or ""
        last = row["assigned_to__last_name"] or ""
        full_name = f"{first} {last}".strip() or row["assigned_to__username"]

        res_times = res_hours_by_tech.get(row["assigned_to"], [])
        avg_hrs = round(sum(res_times) / len(res_times), 2) if res_times else ""

        ws.append(
            [
                full_name,
                row["assigned_to__username"],
                row["total"],
                row["open_count"],
                row["resolved_count"],
                row["closed_count"],
                row["pending_count"],
                row["escalated_count"],
                avg_hrs,
            ]
        )

    _auto_width(ws)


def _sheet_facility_health(ws, qs) -> None:
    headers = [
        "Category",
        "Facility Type",
        "Ticket Count",
        "Open",
        "Resolved / Closed",
    ]
    ws.append(headers)
    _style_headers(ws, len(headers))
    ws.freeze_panes = "A2"

    ACTIVE = ["open", "assigned", "in_progress", "pending"]
    DONE = ["resolved", "closed"]

    # Group by service_item__category and facility_type from location (if present)
    rows = (
        qs.values(
            "service_item__category__name",
            "location__facility_type__name",
        )
        .annotate(
            total=Count("id"),
            open_count=Count("id", filter=Q(status__in=ACTIVE)),
            done_count=Count("id", filter=Q(status__in=DONE)),
        )
        .order_by("-total")
    )

    for row in rows:
        ws.append(
            [
                row["service_item__category__name"] or "Unknown",
                row["location__facility_type__name"] or "—",
                row["total"],
                row["open_count"],
                row["done_count"],
            ]
        )

    _auto_width(ws)


def _sheet_pending_analysis(ws, qs) -> None:
    headers = [
        "Ticket No",
        "Priority",
        "Current Level",
        "Section",
        "Campus",
        "Assigned To",
        "Created At",
        "Paused Since",
        "Total Paused (mins)",
        "Description (100 chars)",
    ]
    ws.append(headers)
    _style_headers(ws, len(headers))
    ws.freeze_panes = "A2"

    pending = qs.filter(status="pending").select_related(
        "priority",
        "section__section_type",
        "section__campus_department__campus",
        "assigned_to",
    )

    for t in pending:
        pause_mins = (
            int(t.accumulated_pause.total_seconds() / 60) if t.accumulated_pause else 0
        )
        ws.append(
            [
                t.ticket_no,
                t.priority.name if t.priority_id else "",
                t.get_current_level_display(),
                (
                    t.section.section_type.name
                    if t.section_id and t.section.section_type_id
                    else ""
                ),
                t.section.campus_department.campus.name if t.section_id else "",
                (
                    (t.assigned_to.get_full_name() or t.assigned_to.username)
                    if t.assigned_to_id
                    else "Unassigned"
                ),
                _fmt(t.created_at),
                _fmt(t.paused_at),
                pause_mins,
                (t.description or "")[:100],
            ]
        )

    _auto_width(ws)


# ── Report type metadata ───────────────────────────────────────────────────────

REPORT_TYPES = [
    {
        "id": "ticket-lifecycle",
        "name": "Ticket Lifecycle Report",
        "description": "Complete ticket audit trail with all lifecycle data including pending reasons",
        "filters": ["date_range", "section", "status"],
        "columns": [
            "ticket_no",
            "status",
            "priority",
            "section",
            "assigned_to",
            "created_at",
            "resolved_at",
        ],
    },
    {
        "id": "technician-performance",
        "name": "Technician Performance Report",
        "description": "Detailed performance metrics for all technicians",
        "filters": ["date_range", "section", "technician"],
        "columns": [
            "technician",
            "total_assigned",
            "resolved",
            "open",
            "avg_resolution_hours",
        ],
    },
    {
        "id": "facility-health",
        "name": "Facility Health Report",
        "description": "Health metrics and maintenance needs by facility type and service category",
        "filters": ["date_range"],
        "columns": ["category", "facility_type", "ticket_count", "open", "resolved"],
    },
    {
        "id": "pending-analysis",
        "name": "Pending Tickets Analysis",
        "description": "All pending tickets with pause durations and priorities",
        "filters": ["section"],
        "columns": [
            "ticket_no",
            "priority",
            "section",
            "paused_since",
            "total_paused_mins",
        ],
    },
    {
        "id": "comprehensive",
        "name": "Comprehensive Report",
        "description": "All metrics combined in a single Excel workbook with multiple sheets",
        "filters": ["date_range"],
        "columns": ["all"],
    },
]

TIMEFRAME_OPTIONS = [
    {"value": "all", "label": "All Time"},
    {"value": "day", "label": "Last 24 Hours"},
    {"value": "week", "label": "Last 7 Days"},
    {"value": "month", "label": "Last 30 Days"},
    {"value": "quarter", "label": "Last 3 Months"},
    {"value": "year", "label": "Last Year"},
]


# ── Views ──────────────────────────────────────────────────────────────────────


class ReportTypesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(
            {
                "report_types": REPORT_TYPES,
                "timeframe_options": TIMEFRAME_OPTIONS,
            }
        )


class GenerateReportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def _get_role(self, request):
        return resolve_role(request)

    def get(self, request):
        report_type = request.query_params.get("report_type", "")
        if not report_type:
            return Response({"error": "report_type is required"}, status=400)

        valid_types = {r["id"] for r in REPORT_TYPES}
        if report_type not in valid_types:
            return Response(
                {"error": f"Unknown report_type: {report_type}"}, status=400
            )

        role = self._get_role(request)
        # scoped_qs: unfiltered by date (aggregate() handles its own window)
        scoped_qs = (
            scoped_ticket_qs(request.user, role)
            if role
            else Ticket.objects.filter(raised_by=request.user)
        )
        qs = _base_qs(request, role)  # date + optional section/tech filters

        wb = openpyxl.Workbook()

        builders = {
            "ticket-lifecycle": _sheet_ticket_lifecycle,
            "technician-performance": _sheet_technician_performance,
            "facility-health": _sheet_facility_health,
            "pending-analysis": _sheet_pending_analysis,
        }

        # Every report gets a Summary sheet first (mirrors analytics overview cards)
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _sheet_summary(ws_summary, scoped_qs, request)

        if report_type == "comprehensive":
            ws1 = wb.create_sheet("Ticket Lifecycle")
            _sheet_ticket_lifecycle(ws1, qs)

            ws2 = wb.create_sheet("Technician Performance")
            _sheet_technician_performance(ws2, qs)

            ws3 = wb.create_sheet("Facility Health")
            _sheet_facility_health(ws3, qs)

            ws4 = wb.create_sheet("Pending Analysis")
            _sheet_pending_analysis(ws4, qs)
        else:
            ws = wb.create_sheet(report_type.replace("-", " ").title())
            builders[report_type](ws, qs)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = report_type.replace("-", "_")
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_report.xlsx"'
        )
        return response
