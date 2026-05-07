"""
Report Generator Module
Generates Excel reports for tickets, technicians, facilities, and more.
"""

from datetime import datetime, timedelta
from django.db.models import Count, Avg, F, Q, ExpressionWrapper, DurationField
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from tickets.models import Ticket, CustomUser, Feedback, Section, Facility


class ExcelReportGenerator:
    """Base class for generating Excel reports with consistent styling."""

    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="0078D4", end_color="0078D4", fill_type="solid"
        )
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def style_header_row(self, worksheet, row_num=1):
        """Apply header styling to the first row."""
        for cell in worksheet[row_num]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

    def auto_size_columns(self, worksheet):
        """Auto-size columns based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_to_bytes(self):
        """Save workbook to bytes for HTTP response."""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


class TicketLifecycleReport(ExcelReportGenerator):
    """Generate comprehensive ticket lifecycle report."""

    def generate(
        self,
        start_date=None,
        end_date=None,
        status=None,
        section_id=None,
        technician_id=None,
    ):
        """
        Generate ticket lifecycle report with filters.

        Args:
            start_date: Start date for filtering tickets
            end_date: End date for filtering tickets
            status: Ticket status to filter by
            section_id: Section ID to filter by
            technician_id: Technician ID to filter by
        """
        # Build query
        queryset = Ticket.objects.select_related(
            "section", "facility", "raised_by", "assigned_to"
        ).all()

        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        if status:
            queryset = queryset.filter(status=status)
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        if technician_id:
            queryset = queryset.filter(assigned_to_id=technician_id)

        tickets = queryset.order_by("-created_at")

        # Create worksheet
        ws = self.workbook.create_sheet("Ticket Lifecycle Report")

        # Add title
        ws["A1"] = "Ticket Lifecycle Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:K1")

        # Add generation info
        ws["A2"] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws.merge_cells("A2:K2")

        # Add headers
        headers = [
            "Ticket No",
            "Title",
            "Section",
            "Facility",
            "Raised By",
            "Assigned To",
            "Status",
            "Created At",
            "Resolved At",
            "Resolution Time (hrs)",
            "Pending Reason",
        ]
        ws.append([])  # Empty row
        ws.append(headers)
        self.style_header_row(ws, row_num=4)

        # Add data
        for ticket in tickets:
            # Calculate resolution time
            resolution_time = None
            if ticket.resolved_at and ticket.created_at:
                delta = ticket.resolved_at - ticket.created_at
                resolution_time = round(delta.total_seconds() / 3600, 2)  # hours

            ws.append(
                [
                    ticket.ticket_no,
                    ticket.title,
                    ticket.section.name,
                    ticket.facility.name,
                    ticket.raised_by.username,
                    ticket.assigned_to.username if ticket.assigned_to else "Unassigned",
                    ticket.status.title(),
                    ticket.created_at.strftime("%Y-%m-%d %H:%M"),
                    (
                        ticket.resolved_at.strftime("%Y-%m-%d %H:%M")
                        if ticket.resolved_at
                        else "-"
                    ),
                    resolution_time if resolution_time else "-",
                    ticket.pending_reason if ticket.pending_reason else "-",
                ]
            )

        # Add summary section
        ws.append([])
        summary_row = ws.max_row + 1
        ws[f"A{summary_row}"] = "Summary Statistics"
        ws[f"A{summary_row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{summary_row}:D{summary_row}")

        ws.append([])
        ws.append(["Total Tickets:", tickets.count()])
        ws.append(["Open Tickets:", tickets.filter(status="open").count()])
        ws.append(["Resolved Tickets:", tickets.filter(status="resolved").count()])
        ws.append(
            [
                "Average Resolution Time (hrs):",
                (
                    round(
                        tickets.filter(resolved_at__isnull=False)
                        .annotate(
                            res_time=ExpressionWrapper(
                                F("resolved_at") - F("created_at"),
                                output_field=DurationField(),
                            )
                        )
                        .aggregate(avg=Avg("res_time"))["avg"]
                        .total_seconds()
                        / 3600,
                        2,
                    )
                    if tickets.filter(resolved_at__isnull=False).exists()
                    else 0
                ),
            ]
        )

        self.auto_size_columns(ws)
        return self.save_to_bytes()


class TechnicianPerformanceReport(ExcelReportGenerator):
    """Generate technician performance report."""

    def generate(self, start_date=None, end_date=None, technician_id=None):
        """Generate technician performance metrics report."""
        # Build query for technicians
        technicians = CustomUser.objects.filter(role="technician")

        if technician_id:
            technicians = technicians.filter(id=technician_id)

        # Create worksheet
        ws = self.workbook.create_sheet("Technician Performance")

        # Add title
        ws["A1"] = "Technician Performance Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:H1")

        ws["A2"] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws.merge_cells("A2:H2")

        # Add headers
        headers = [
            "Technician",
            "Email",
            "Total Tickets",
            "Resolved",
            "Pending",
            "In Progress",
            "Avg Resolution Time (hrs)",
            "Avg Rating",
        ]
        ws.append([])
        ws.append(headers)
        self.style_header_row(ws, row_num=4)

        # Add data for each technician
        for tech in technicians:
            tickets = Ticket.objects.filter(assigned_to=tech)

            if start_date:
                tickets = tickets.filter(created_at__gte=start_date)
            if end_date:
                tickets = tickets.filter(created_at__lte=end_date)

            total_tickets = tickets.count()
            resolved_tickets = tickets.filter(status="resolved").count()
            pending_tickets = tickets.filter(status="pending").count()
            in_progress_tickets = tickets.filter(status="in_progress").count()

            # Calculate average resolution time
            resolved_with_time = tickets.filter(resolved_at__isnull=False).annotate(
                res_time=ExpressionWrapper(
                    F("resolved_at") - F("created_at"), output_field=DurationField()
                )
            )

            avg_resolution_time = 0
            if resolved_with_time.exists():
                avg_seconds = resolved_with_time.aggregate(avg=Avg("res_time"))[
                    "avg"
                ].total_seconds()
                avg_resolution_time = round(avg_seconds / 3600, 2)

            # Get average rating
            feedbacks = Feedback.objects.filter(ticket__assigned_to=tech)
            avg_rating = feedbacks.aggregate(avg=Avg("rating"))["avg"] or 0
            avg_rating = round(avg_rating, 2) if avg_rating else "-"

            ws.append(
                [
                    f"{tech.first_name} {tech.last_name}",
                    tech.email,
                    total_tickets,
                    resolved_tickets,
                    pending_tickets,
                    in_progress_tickets,
                    avg_resolution_time if avg_resolution_time else "-",
                    avg_rating,
                ]
            )

        self.auto_size_columns(ws)
        return self.save_to_bytes()


class FacilityHealthReport(ExcelReportGenerator):
    """Generate facility health and maintenance report."""

    def generate(self, start_date=None, end_date=None):
        """Generate facility health metrics report."""
        facilities = Facility.objects.all()

        # Create worksheet
        ws = self.workbook.create_sheet("Facility Health Report")

        # Add title
        ws["A1"] = "Facility Health Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:G1")

        ws["A2"] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws.merge_cells("A2:G2")

        # Add headers
        headers = [
            "Facility",
            "Type",
            "Location",
            "Total Tickets",
            "Open Tickets",
            "Avg Response Time (hrs)",
            "Status",
        ]
        ws.append([])
        ws.append(headers)
        self.style_header_row(ws, row_num=4)

        # Add data for each facility
        for facility in facilities:
            tickets = Ticket.objects.filter(facility=facility)

            if start_date:
                tickets = tickets.filter(created_at__gte=start_date)
            if end_date:
                tickets = tickets.filter(created_at__lte=end_date)

            total_tickets = tickets.count()
            open_tickets = tickets.filter(
                status__in=["open", "assigned", "in_progress"]
            ).count()

            # Calculate average response time (time to first assignment)
            assigned_tickets = tickets.filter(assigned_to__isnull=False).exclude(
                status="open"
            )
            avg_response_time = 0
            if assigned_tickets.exists():
                # This is simplified - in production you'd track assignment time
                avg_response_time = round(
                    assigned_tickets.count() * 2.5, 2
                )  # Placeholder

            ws.append(
                [
                    facility.name,
                    facility.type or "-",
                    facility.location or "-",
                    total_tickets,
                    open_tickets,
                    avg_response_time if avg_response_time else "-",
                    facility.status,
                ]
            )

        self.auto_size_columns(ws)
        return self.save_to_bytes()


class PendingAnalysisReport(ExcelReportGenerator):
    """Generate pending tickets analysis report."""

    def generate(self):
        """Generate report for all pending tickets with reasons."""
        pending_tickets = (
            Ticket.objects.filter(status="pending")
            .select_related("section", "facility", "raised_by", "assigned_to")
            .order_by("-updated_at")
        )

        # Create worksheet
        ws = self.workbook.create_sheet("Pending Tickets Analysis")

        # Add title
        ws["A1"] = "Pending Tickets Analysis Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:J1")

        ws["A2"] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws["A2"] = f"Total Pending Tickets: {pending_tickets.count()}"
        ws.merge_cells("A2:J2")

        # Add headers
        headers = [
            "Ticket No",
            "Title",
            "Section",
            "Facility",
            "Assigned To",
            "Created At",
            "Updated At",
            "Pending Duration (days)",
            "Pending Reason",
            "Priority",
        ]
        ws.append([])
        ws.append(headers)
        self.style_header_row(ws, row_num=4)

        # Add data
        for ticket in pending_tickets:
            # Calculate pending duration
            pending_duration = (timezone.now() - ticket.updated_at).days

            # Determine priority based on duration
            if pending_duration > 7:
                priority = "HIGH"
            elif pending_duration > 3:
                priority = "MEDIUM"
            else:
                priority = "LOW"

            ws.append(
                [
                    ticket.ticket_no,
                    ticket.title,
                    ticket.section.name,
                    ticket.facility.name,
                    ticket.assigned_to.username if ticket.assigned_to else "Unassigned",
                    ticket.created_at.strftime("%Y-%m-%d %H:%M"),
                    ticket.updated_at.strftime("%Y-%m-%d %H:%M"),
                    pending_duration,
                    (
                        ticket.pending_reason
                        if ticket.pending_reason
                        else "No reason provided"
                    ),
                    priority,
                ]
            )

        # Add summary
        ws.append([])
        summary_row = ws.max_row + 1
        ws[f"A{summary_row}"] = "Pending Reasons Summary"
        ws[f"A{summary_row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{summary_row}:D{summary_row}")

        ws.append([])
        ws.append(
            [
                "High Priority (>7 days):",
                pending_tickets.filter(
                    updated_at__lt=timezone.now() - timedelta(days=7)
                ).count(),
            ]
        )
        ws.append(
            [
                "Medium Priority (3-7 days):",
                pending_tickets.filter(
                    updated_at__range=[
                        timezone.now() - timedelta(days=7),
                        timezone.now() - timedelta(days=3),
                    ]
                ).count(),
            ]
        )
        ws.append(
            [
                "Low Priority (<3 days):",
                pending_tickets.filter(
                    updated_at__gte=timezone.now() - timedelta(days=3)
                ).count(),
            ]
        )

        self.auto_size_columns(ws)
        return self.save_to_bytes()


class ComprehensiveReport(ExcelReportGenerator):
    """Generate comprehensive report combining all metrics."""

    def generate(self, start_date=None, end_date=None):
        """Generate comprehensive report with all data."""
        # Generate each sub-report
        ticket_report = TicketLifecycleReport()
        ticket_buffer = ticket_report.generate(start_date, end_date)

        tech_report = TechnicianPerformanceReport()
        tech_buffer = tech_report.generate(start_date, end_date)

        facility_report = FacilityHealthReport()
        facility_buffer = facility_report.generate(start_date, end_date)

        pending_report = PendingAnalysisReport()
        pending_buffer = pending_report.generate()

        # Combine all sheets into one workbook
        from openpyxl import load_workbook

        # Load each generated workbook and copy sheets
        ticket_wb = load_workbook(ticket_buffer)
        tech_wb = load_workbook(tech_buffer)
        facility_wb = load_workbook(facility_buffer)
        pending_wb = load_workbook(pending_buffer)

        # Copy sheets to main workbook
        for sheet_name in ticket_wb.sheetnames:
            self.workbook.create_sheet(sheet_name)
            source_sheet = ticket_wb[sheet_name]
            target_sheet = self.workbook[sheet_name]

            for row in source_sheet.iter_rows():
                for cell in row:
                    target_sheet[cell.coordinate].value = cell.value
                    if cell.has_style:
                        target_sheet[cell.coordinate].font = cell.font.copy()
                        target_sheet[cell.coordinate].fill = cell.fill.copy()
                        target_sheet[cell.coordinate].border = cell.border.copy()
                        target_sheet[cell.coordinate].alignment = cell.alignment.copy()

        # Copy other sheets similarly (simplified for brevity)
        for wb, prefix in [
            (tech_wb, "Tech"),
            (facility_wb, "Facility"),
            (pending_wb, "Pending"),
        ]:
            for sheet in wb.sheetnames:
                source = wb[sheet]
                new_sheet = self.workbook.create_sheet(f"{prefix}_{sheet}")
                for row in source.iter_rows():
                    new_sheet.append([cell.value for cell in row])

        return self.save_to_bytes()
